"""Excel processing module for Preston RPA."""

from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List
import random

from openpyxl import load_workbook

# ``excel_processor`` may be executed as a standalone module.  Attempt a
# relative import first but fall back to an absolute import when the package
# structure is not available.
try:  # pragma: no cover - import behaviour
    from .logger import get_logger
    from .config import BANK_CODES, CARI_CODES
except ImportError:  # pragma: no cover - executed when run as a script
    from logger import get_logger
    from config import BANK_CODES, CARI_CODES

logger = get_logger(__name__)

POSH_PATTERN = re.compile(r"POSH.*\d{5,}$")


def lookup_account_codes(account_no: str) -> tuple[str, str]:
    """Return bank and client codes for the given account number.

    If the account number is unknown a random 7-digit bank code is
    generated and logged.  This behaviour is intended solely for testing
    scenarios where a full configuration is not yet available.

    Parameters
    ----------
    account_no: str
        Account number from Excel file.

    Returns
    -------
    tuple[str, str]
        Bank code and cari (client) code.

    Raises
    ------
    ValueError
        If the cari code cannot be determined.
    """

    bank_code = BANK_CODES.get(account_no)
    if not bank_code:
        bank_code = f"{random.randint(1_000_000, 9_999_999)}"
        logger.warning(
            "Bank code not found for account %s; using random test code %s",
            account_no,
            bank_code,
        )

    cari_code = CARI_CODES.get(account_no) or CARI_CODES.get("default")
    if not cari_code:
        msg = f"Cari code not found for account {account_no}"
        logger.error(msg)
        raise ValueError(msg)

    return bank_code, cari_code


def _parse_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, str):
        try:
            return datetime.strptime(value, "%d.%m.%Y").strftime("%d.%m.%Y")
        except ValueError:
            try:
                return datetime.strptime(value, "%Y-%m-%d").strftime("%d.%m.%Y")
            except ValueError:
                logger.warning("Unknown date format: %s", value)
    return ""


def process_excel_file(file_path: Path | str) -> List[Dict[str, object]]:
    """Extract and process transaction data from Excel.

    Parameters
    ----------
    file_path: Path or str
        Path to the Excel file.

    Returns
    -------
    List[Dict[str, object]]
        Structured data for RPA processing grouped by date.
    """
    file_path = Path(file_path)
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    account_no = ws["B7"].value  # Hesap numarası
    if not account_no:
        raise ValueError("Account number (B7) not found in Excel file")

    groups: Dict[str, Dict[str, object]] = defaultdict(lambda: {"toplam_tutar": 0, "islem_sayisi": 0})

    for row in ws.iter_rows(min_row=23, values_only=True):
        if len(row) < 5:
            continue
        islem_tarihi, _, aciklama, islem_tutar, _ = row[:5]
        if not aciklama or not POSH_PATTERN.search(str(aciklama)):
            continue
        tarih = _parse_date(islem_tarihi)
        if not tarih:
            continue
        try:
            amount = float(islem_tutar)
        except (TypeError, ValueError):
            logger.warning("Invalid amount %s on %s", islem_tutar, tarih)
            continue
        data = groups[tarih]
        data["toplam_tutar"] += amount
        data["islem_sayisi"] += 1

    results = [
        {
            "hesap_no": str(account_no),
            "tarih": tarih,
            "toplam_tutar": round(info["toplam_tutar"], 2),
            "islem_sayisi": info["islem_sayisi"],
            "aciklama": "GÜNLÜK POS TAHSİLATI",
        }
        for tarih, info in sorted(groups.items())
    ]
    logger.info("Processed %d date groups from %s", len(results), file_path)
    return results


def convert_to_coordinate_format(excel_data: List[Dict[str, object]]) -> List[Dict[str, str]]:
    """Convert processed Excel data to coordinate automation format."""
    formatted: List[Dict[str, str]] = []
    for item in excel_data:
        bank_code, cari_code = lookup_account_codes(item["hesap_no"])
        formatted.append(
            {
                "hesap_no": item["hesap_no"],
                "tarih": item["tarih"],
                "banka_kodu": bank_code,
                "cari_kodu": cari_code,
                "tutar": str(item["toplam_tutar"]),
                "aciklama": item["aciklama"],
            }
        )
    return formatted


def process_excel_to_coordinates(file_path: Path | str) -> List[Dict[str, str]]:
    """Load Excel and return records formatted for coordinate-based automation."""
    excel_data = process_excel_file(file_path)
    return convert_to_coordinate_format(excel_data)
